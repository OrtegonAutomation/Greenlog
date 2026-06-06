from __future__ import annotations

import json
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Exceles de Referencia" / "Nuevos Archivos de referencia" / "Plan de Monitoreos 2026.xlsx"
OUT = ROOT / "src" / "services" / "monitoreos_data.json"
SHEET = "Tabla total  2026"

MONTH_INDEX = {
    "ENERO": 0,
    "FEBRERO": 1,
    "MARZO": 2,
    "ABRIL": 3,
    "MAYO": 4,
    "JUNIO": 5,
    "JULIO": 6,
    "AGOSTO": 7,
    "SEPTIEMBRE": 8,
    "SETIEMBRE": 8,
    "OCTUBRE": 9,
    "NOVIEMBRE": 10,
    "DICIEMBRE": 11,
}

CANONICAL_MATRICES = {
    "ARD": "ARD",
    "ARND": "ARnD",
    "AGUA SUPERFICIAL": "Agua Superficial",
    "AGUA SUBTERRANEA": "Agua Subterranea",
    "AGUA MARINA": "Agua Marina",
    "AIRE": "Aire",
    "SUELO": "Suelo",
    "ISOCINETICO": "Isocinetico",
    "OLORES OFENSIVOS": "Olores Ofensivos",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def number(value: Any) -> float:
    if value is None or value == "":
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def normalize_key(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    normalized = "".join(ch if ch.isalnum() else " " for ch in normalized.upper())
    return " ".join(normalized.split())


def canonical_matriz(value: str) -> str:
    key = normalize_key(value)
    return CANONICAL_MATRICES.get(key, clean(value))


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb[SHEET]
    headers = [clean(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: i for i, name in enumerate(headers)}

    grouped: dict[tuple[str, ...], dict[str, Any]] = {}
    estaciones_por_zona: dict[str, set[str]] = defaultdict(set)
    canonical_seen: dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        zona = clean(row[index["Zona"]])
        estacion = clean(row[index["Estación / sistio"]])
        raw_matriz = clean(row[index["Matriz"]])
        matriz = canonical_matriz(raw_matriz)
        matriz_key = normalize_key(matriz)
        if matriz_key in canonical_seen:
            matriz = canonical_seen[matriz_key]
        else:
            canonical_seen[matriz_key] = matriz
        parametro = clean(row[index["Parametro"]])
        if not zona or not estacion or not matriz or not parametro:
            continue

        permiso = clean(row[index["Permiso"]])
        receptor = clean(row[index["Receptor /uso"]])
        requerimiento = clean(row[index["Requerimiento establecido por"]])
        norma = clean(row[index["Norma referencia"]])
        item = clean(row[index["Item"]])
        sistema = clean(row[index["Sistema"]])
        chemilab = number(row[index["Chemilab"]])
        puntos = number(row[index["Puntos/Q"]])
        compuesto = number(row[index["Compuesto"]])
        total_chemilab = number(row[index["Total Chemilab"]])
        mes = clean(row[index["Mes"]])

        key = (
            zona,
            estacion,
            matriz,
            permiso,
            receptor,
            requerimiento,
            norma,
            parametro,
            item,
            sistema,
        )

        if key not in grouped:
            grouped[key] = {
                "zona": zona,
                "estacion": estacion,
                "matriz": matriz,
                "permiso": permiso,
                "receptor": receptor,
                "requerimiento": requerimiento,
                "norma": norma,
                "parametro": parametro,
                "item": item,
                "sistema": sistema,
                "chemilab": chemilab,
                "puntos": puntos,
                "compuesto": compuesto,
                "mes": mes,
                "totalChemilab": total_chemilab,
            }
        else:
            current = grouped[key]
            current["chemilab"] = current["chemilab"] or chemilab
            current["puntos"] = current["puntos"] or puntos
            current["compuesto"] = current["compuesto"] or compuesto
            current["mes"] = current["mes"] or mes
            current["totalChemilab"] = current["totalChemilab"] or total_chemilab

        month_key = normalize_key(mes)
        if month_key in MONTH_INDEX and chemilab:
            prices = grouped[key].setdefault("preciosMensuales", {})
            prices[str(MONTH_INDEX[month_key])] = chemilab

        estaciones_por_zona[zona].add(estacion)

    matriz_data = sorted(
        grouped.values(),
        key=lambda r: (
            r["zona"],
            r["estacion"],
            normalize_key(r["matriz"]),
            normalize_key(r["parametro"]),
        ),
    )
    for row in matriz_data:
        if "preciosMensuales" in row:
            row["preciosMensuales"] = {
                int(k): v for k, v in sorted(row["preciosMensuales"].items(), key=lambda kv: int(kv[0]))
            }

    payload = {
        "estacionesPorZona": {
            zona: sorted(estaciones)
            for zona, estaciones in sorted(estaciones_por_zona.items())
        },
        "matrizData": matriz_data,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {len(matriz_data)} rows to {OUT}")


if __name__ == "__main__":
    main()
