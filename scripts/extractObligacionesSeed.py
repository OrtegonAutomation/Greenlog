"""
Extrae las obligaciones reales de la Matriz Provisiones 2 a un módulo TypeScript
que se commitea como seed: src/data/obligacionesSeed.ts

Uso (1 sola vez, o cuando cambie la matriz):
    python scripts/extractObligacionesSeed.py
"""
import openpyxl
import json
import unicodedata
from pathlib import Path

ROOT = Path(__file__).parent.parent
SRC = ROOT / 'Exceles de Referencia' / 'Provisiones' / \
    'RE_ Plan de Acción y Solicitud Provisión Ambiental CAR - RES 06257000283 de 2025 - Otorga concesión de aguas superficiales - Estación Albán' / \
    'Matriz Provisiones 2.xlsx'
DEST = ROOT / 'src' / 'data' / 'obligacionesSeed.ts'

# Columnas (1-indexed) en la matriz:
# 1=ID, 2=ZONA, 3=SISTEMA, 4=Sector/Proyecto, 5=PERMISO,
# 6=Autoridad Ambiental, 7=Numero expediente, 8=Tipo Acto, 9=Numero Acto,
# 10=Fecha Acto, 11=CATEGORIA, 12=RESPONSABLE,
# 47=CAPEX/OPEX, 61=SALDO DISPONIBLE 2026 DESPUÉS DE USO PROYECTADO

def title_loc(s):
    if not s:
        return ''
    s = unicodedata.normalize('NFC', str(s).strip())
    if s.isupper():
        s = ' '.join(w.capitalize() for w in s.lower().split())
    return s

def fmt_date(v):
    if v is None or v == '':
        return ''
    if hasattr(v, 'isoformat'):
        return v.date().isoformat() if hasattr(v, 'date') else v.isoformat()
    return str(v).strip()

def fmt_str(v):
    if v is None:
        return ''
    return str(v).strip()

def fmt_num(v):
    if v is None or v == '':
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0

def main():
    print(f'Leyendo: {SRC}')
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active

    obligaciones = []
    seen_ids = set()
    for r in range(2, ws.max_row + 1):
        idObligacion = fmt_str(ws.cell(row=r, column=1).value)
        if not idObligacion:
            continue
        # Solo IDs con formato XXX_NN (ej. COMP_01, INV_05). Descarta filas
        # de cabeceras de sub-grupos, totales, fórmulas y demás ruido.
        import re
        if not re.match(r'^[A-Z]+_\d+$', idObligacion):
            continue
        # Dedup: conservar la primera ocurrencia
        if idObligacion in seen_ids:
            continue
        seen_ids.add(idObligacion)

        zona = fmt_str(ws.cell(row=r, column=2).value)
        sistema = fmt_str(ws.cell(row=r, column=3).value)
        sector = fmt_str(ws.cell(row=r, column=4).value)
        permiso = fmt_str(ws.cell(row=r, column=5).value)
        autoridadEmite = fmt_str(ws.cell(row=r, column=6).value)
        expediente = fmt_str(ws.cell(row=r, column=7).value)
        actoTipo = fmt_str(ws.cell(row=r, column=8).value)
        actoNumero = fmt_str(ws.cell(row=r, column=9).value)
        actoFecha = fmt_date(ws.cell(row=r, column=10).value)
        categoria = fmt_str(ws.cell(row=r, column=11).value)
        responsable = fmt_str(ws.cell(row=r, column=12).value)
        capexOpex = fmt_str(ws.cell(row=r, column=47).value)
        saldoDisponible = fmt_num(ws.cell(row=r, column=61).value)

        obligaciones.append({
            'idObligacion': idObligacion,
            'zona': zona,
            'sistema': sistema,
            'sector': sector,
            'permiso': permiso,
            'autoridadEmite': autoridadEmite,
            'expediente': expediente,
            'actoTipo': actoTipo,
            'actoNumero': actoNumero,
            'actoFecha': actoFecha,
            'categoria': categoria,
            'responsable': responsable,
            'fuentePresupuesto': capexOpex.upper() if capexOpex.upper() in ('OPEX', 'CAPEX') else 'OPEX',
            'saldoDisponible': round(saldoDisponible, 2),
        })

    print(f'  → {len(obligaciones)} obligaciones extraídas')

    # Generar TS
    lines = [
        '// Auto-generado desde Matriz Provisiones 2.xlsx por scripts/extractObligacionesSeed.py',
        '// NO editar a mano. Para regenerar: `python scripts/extractObligacionesSeed.py`',
        '',
        'export interface ObligacionSeed {',
        '  idObligacion: string;',
        '  zona: string;',
        '  sistema: string;',
        '  sector: string;',
        '  permiso: string;',
        '  autoridadEmite: string;',
        '  expediente: string;',
        '  actoTipo: string;',
        '  actoNumero: string;',
        '  actoFecha: string;  // ISO yyyy-mm-dd',
        '  categoria: string;',
        '  responsable: string;',
        '  fuentePresupuesto: "OPEX" | "CAPEX";',
        '  saldoDisponible: number;  // COP',
        '}',
        '',
        f'export const OBLIGACIONES_SEED: ObligacionSeed[] = {json.dumps(obligaciones, ensure_ascii=False, indent=2)};',
        '',
    ]
    DEST.write_text('\n'.join(lines), encoding='utf-8')
    print(f'  → escrito: {DEST}')

if __name__ == '__main__':
    main()
